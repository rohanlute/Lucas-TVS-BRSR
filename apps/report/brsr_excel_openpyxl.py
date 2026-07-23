# apps/report/brsr_excel_openpyxl.py
"""
Builds the BRSR report as an .xlsx workbook matching the same structure as
the PDF (brsr_pdf_reportlab.py): navy/gold styling, one sheet per Section
A/B and per Principle, with matrix (P1-P9) and table-shaped answers
rendered as real grids -- not flattened into a single "Answer" column.

--------------------------------------------------------------------------
WHY THIS REUSES _flatten_rows FROM brsr_pdf_reportlab.py
--------------------------------------------------------------------------
The previous Excel view built its own flat "No. | Section | Sub-section /
Principle | Question | Type | Answer" sheet directly off report_sections,
writing only `row.get("answer_value", "")` per top-level question -- so
every sub_question (table cells, matrix grids, multi-field questions) was
silently dropped, and nothing here shared logic with the PDF's rendering.

_flatten_rows() in brsr_pdf_reportlab.py is pure data-shaping -- it takes
row dicts and yields ("heading"/"matrix"/"table"/"row", ...) tuples of
plain strings/lists, with no ReportLab Paragraph/Table objects involved.
So it's exactly as reusable here as it is there: import it, walk the same
items, and Excel + PDF can never structurally diverge from each other
again.
"""

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .brsr_pdf_reportlab import _flatten_rows, _display_financial_year, _short_company_name

# ---------------------------------------------------------------------------
# Palette -- matches brsr_pdf_reportlab.py's NAVY/GOLD_BORDER/etc.
# ---------------------------------------------------------------------------
NAVY = "FF1B2A56"
GOLD = "FFF4B000"
WHITE = "FFFFFFFF"
ORANGE = "FFF4A300"
BODY = "FF1A1A1A"
MUTED = "FF6B7280"

NAVY_FILL = PatternFill("solid", fgColor=NAVY)
THIN = Side(style="thin", color=GOLD)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FONT_WHITE_BOLD = Font(color=WHITE, bold=True, size=10)
FONT_BLACK_BOLD = Font(color=BODY, bold=True, size=10)
FONT_BLACK = Font(color=BODY, size=9)
FONT_MUTED_ITALIC = Font(color=MUTED, italic=True, size=9)
FONT_ORANGE_BOLD = Font(color=ORANGE, bold=True, size=11)
FONT_TITLE = Font(color=WHITE, bold=True, size=13)
FONT_INDICATOR = Font(color=MUTED, bold=True, size=10)

WRAP_TOP_LEFT = Alignment(wrap_text=True, vertical="top", horizontal="left")
WRAP_CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")
CENTER_MIDDLE = Alignment(vertical="center", horizontal="left", indent=1)

VALUE_COLS = 9  # matches the P1..P9 matrix width; generic tables also budget within this
NUMBER_COL_WIDTH = 6
LABEL_COL_WIDTH = 46
VALUE_COL_WIDTH = 13
LAST_COL = 2 + VALUE_COLS  # column K


def _setup_sheet(ws):
    ws.column_dimensions["A"].width = NUMBER_COL_WIDTH
    ws.column_dimensions["B"].width = LABEL_COL_WIDTH
    for i in range(3, 3 + VALUE_COLS):
        ws.column_dimensions[get_column_letter(i)].width = VALUE_COL_WIDTH
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"


def _write_full_width(ws, row, text, font, height=None):
    ws.cell(row=row, column=1, value=text)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=LAST_COL)
    cell = ws.cell(row=row, column=1)
    cell.font = font
    cell.alignment = CENTER_MIDDLE
    if height:
        ws.row_dimensions[row].height = height
    return row + 1


def _write_section_title(ws, row, text):
    row_next = _write_full_width(ws, row, text, FONT_TITLE, height=22)
    ws.cell(row=row, column=1).fill = NAVY_FILL
    for col in range(1, LAST_COL + 1):
        ws.cell(row=row, column=col).fill = NAVY_FILL
    return row_next


def _write_subheading(ws, row, text):
    return _write_full_width(ws, row, text, FONT_ORANGE_BOLD)


def _write_indicator_label(ws, row, text):
    return _write_full_width(ws, row, text, FONT_INDICATOR)


def _write_heading_item(ws, row, number, text):
    return _write_full_width(ws, row, f"{number}.     {text}", FONT_ORANGE_BOLD)


def _fmt_bool(value):
    if value is True or value == "True":
        return "Yes"
    if value is False or value == "False":
        return ""
    return value


def _write_plain_row(ws, row, number, text, answer):
    num_cell = ws.cell(row=row, column=1, value=f"{number}." if number else "")
    num_cell.font = FONT_BLACK_BOLD
    num_cell.alignment = Alignment(horizontal="right", vertical="top")
    num_cell.border = BORDER

    q_cell = ws.cell(row=row, column=2, value=text or "")
    q_cell.font = FONT_WHITE_BOLD
    q_cell.fill = NAVY_FILL
    q_cell.alignment = WRAP_TOP_LEFT
    q_cell.border = BORDER

    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=LAST_COL)
    a_cell = ws.cell(row=row, column=3, value=answer if answer else "-")
    a_cell.font = FONT_BLACK if answer else FONT_MUTED_ITALIC
    a_cell.alignment = WRAP_TOP_LEFT
    for col in range(3, LAST_COL + 1):
        ws.cell(row=row, column=col).border = BORDER

    ws.row_dimensions[row].height = 26
    return row + 1


def _write_matrix_grid(ws, row, label, columns, matrix_rows):
    if label:
        row = _write_subheading(ws, row, label)

    for c in (1, 2):
        cell = ws.cell(row=row, column=c, value="")
        cell.fill = NAVY_FILL
        cell.border = BORDER
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    for i, col_name in enumerate(columns):
        c = ws.cell(row=row, column=3 + i, value=col_name)
        c.font = FONT_WHITE_BOLD
        c.fill = NAVY_FILL
        c.alignment = WRAP_CENTER
        c.border = BORDER
    row += 1

    for mrow in matrix_rows:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        for c in (1, 2):
            cell = ws.cell(row=row, column=c)
            cell.fill = NAVY_FILL
            cell.border = BORDER
        label_cell = ws.cell(row=row, column=1, value=mrow.get("label", ""))
        label_cell.font = FONT_WHITE_BOLD
        label_cell.alignment = WRAP_TOP_LEFT

        values = mrow.get("values", {})
        for i, col_name in enumerate(columns):
            val = _fmt_bool(values.get(col_name, ""))
            v_cell = ws.cell(row=row, column=3 + i, value=val if val not in (None, "") else "-")
            v_cell.font = FONT_BLACK
            v_cell.alignment = WRAP_CENTER
            v_cell.border = BORDER
        ws.row_dimensions[row].height = 22
        row += 1

    return row + 1  # spacer after the grid


def _write_table_grid(ws, row, question_text, table_headers, table_rows):
    if question_text:
        row = _write_subheading(ws, row, question_text)

    if not table_headers or not table_rows:
        return _write_plain_row(ws, row, "", "", "")

    for header_row in table_headers:
        for i, h in enumerate(header_row):
            c = ws.cell(row=row, column=1 + i, value=h or "")
            c.font = FONT_WHITE_BOLD
            c.fill = NAVY_FILL
            c.alignment = WRAP_CENTER
            c.border = BORDER
        row += 1

    for data_row in table_rows:
        for i, val in enumerate(data_row):
            display_val = val if val not in (None, "") else ("" if i == 0 else "-")
            c = ws.cell(row=row, column=1 + i, value=_fmt_bool(display_val))
            if i == 0:
                c.font = FONT_WHITE_BOLD
                c.fill = NAVY_FILL
                c.alignment = WRAP_TOP_LEFT
            else:
                c.font = FONT_BLACK
                c.alignment = WRAP_CENTER
            c.border = BORDER
        ws.row_dimensions[row].height = 22
        row += 1

    return row + 1


def _write_rows_block(ws, row, rows):
    """Walks _flatten_rows(rows) -- the same normalized items the PDF
    renders from -- and writes each item in its proper grid/row form."""
    for item in _flatten_rows(rows):
        kind = item[0]
        if kind == "heading":
            _, number, text, _, _ = item
            row = _write_heading_item(ws, row, number, text)
        elif kind == "matrix":
            _, label, columns, matrix_rows, _ = item
            row = _write_matrix_grid(ws, row, label, columns, matrix_rows)
        elif kind == "table":
            _, question_text, table_rows, table_headers, _ = item
            row = _write_table_grid(ws, row, question_text, table_headers, table_rows)
        else:
            _, number, text, q_type, answer = item
            if q_type == "checkbox" and isinstance(answer, str) and answer:
                low = answer.lower()
                if low in ("true", "yes", "1", "on"):
                    answer = "Yes"
                elif low in ("false", "no", "0", "off"):
                    answer = ""
            row = _write_plain_row(ws, row, number, text, answer)
    return row


def _safe_sheet_title(title):
    """Excel sheet names: max 31 chars, no []:\\/?* characters."""
    for ch in "[]:\\/?*":
        title = title.replace(ch, "")
    return title[:31] or "Sheet"


def generate_brsr_excel(financial_year=None, assignment_id=None, company_name="Lucas TVS Ltd", company_cin=""):
    from .brsr_report_data import get_brsr_report_data

    report_sections = get_brsr_report_data(financial_year, assignment_id)

    wb = Workbook()
    wb.remove(wb.active)

    # --- Overview sheet -----------------------------------------------
    ov = wb.create_sheet("Overview")
    _setup_sheet(ov)
    ov.cell(row=1, column=1, value="Business Responsibility & Sustainability Report").font = Font(bold=True, size=15, color="FF1B2A56")
    ov.cell(row=2, column=1, value=_short_company_name(company_name)).font = Font(bold=True, size=12)
    if company_cin:
        ov.cell(row=3, column=1, value=f"CIN: {company_cin}").font = Font(size=10)
    ov.cell(row=4, column=1, value=_display_financial_year(financial_year)).font = Font(size=11, bold=True, color="FF1B2A56")

    section_titles = {"section_a": "Section A", "section_b": "Section B"}

    for block in report_sections:
        section = block.get("section")
        section_name = getattr(section, "name", "Section")
        section_code = getattr(section, "code", "")

        if not block.get("is_principle_section"):
            sheet_title = section_titles.get(section_code, _safe_sheet_title(section_name))
            ws = wb.create_sheet(_safe_sheet_title(sheet_title))
            _setup_sheet(ws)
            row = _write_section_title(ws, 1, section_name.upper())
            for sub in block.get("sub_sections", []):
                sub_title = sub.get("title", "General")
                if sub_title != "General" and sub_title:
                    row = _write_subheading(ws, row, sub_title)
                if sub.get("rows"):
                    row = _write_rows_block(ws, row, sub["rows"])
        else:
            for p_block in block.get("principle_blocks", []):
                principle = p_block.get("principle")
                if hasattr(principle, "principle_number"):
                    p_num = principle.principle_number
                    p_title = getattr(principle, "title", "")
                else:
                    p_num = principle.get("principle_number", "")
                    p_title = principle.get("title", "")

                ws = wb.create_sheet(_safe_sheet_title(f"P{p_num}"))
                _setup_sheet(ws)
                row = _write_section_title(ws, 1, f"Principle {p_num}: {p_title}")

                rows_ = p_block.get("rows", [])
                essential = [r for r in rows_ if (r.get("sub_section") or "").lower().startswith("essential")]
                leadership = [r for r in rows_ if (r.get("sub_section") or "").lower().startswith("leadership")]
                other = [r for r in rows_ if r not in essential and r not in leadership]

                if essential:
                    row = _write_indicator_label(ws, row, "Essential Indicators")
                    row = _write_rows_block(ws, row, essential)
                if leadership:
                    row = _write_indicator_label(ws, row, "Leadership Indicators")
                    row = _write_rows_block(ws, row, leadership)
                if other:
                    row = _write_rows_block(ws, row, other)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


__all__ = ["generate_brsr_excel"]