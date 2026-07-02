from django.core.management.base import BaseCommand
from openpyxl import load_workbook

from apps.companies.models import Country, State, City


class Command(BaseCommand):
    help = "Import Countries, States and Cities"

    def add_arguments(self, parser):
        parser.add_argument(
            "excel_file",
            type=str,
            help="Path to Excel File"
        )

    def handle(self, *args, **options):

        workbook = load_workbook(options["excel_file"])

        ###########################
        # Countries
        ###########################

        sheet = workbook["Countries"]

        for row in sheet.iter_rows(min_row=2, values_only=True):

            name, iso = row

            Country.objects.update_or_create(
                iso_code=iso,
                defaults={
                    "name": name
                }
            )

        self.stdout.write(
            self.style.SUCCESS("Countries Imported")
        )

        ###########################
        # States
        ###########################

        sheet = workbook["States"]

        for row in sheet.iter_rows(min_row=2, values_only=True):

            country_iso, state_name, state_code = row

            try:

                country = Country.objects.get(
                    iso_code=country_iso
                )

                State.objects.update_or_create(
                    country=country,
                    state_code=state_code,
                    defaults={
                        "name": state_name
                    }
                )

            except Country.DoesNotExist:

                self.stdout.write(
                    self.style.ERROR(
                        f"Country {country_iso} not found."
                    )
                )

        self.stdout.write(
            self.style.SUCCESS("States Imported")
        )

        ###########################
        # Cities
        ###########################

        sheet = workbook["Cities"]

        for row in sheet.iter_rows(min_row=2, values_only=True):

            country_iso, state_code, city_name = row

            try:

                country = Country.objects.get(
                    iso_code=country_iso
                )

                state = State.objects.get(
                    country=country,
                    state_code=state_code
                )

                City.objects.update_or_create(
                    country=country,
                    state=state,
                    name=city_name
                )

            except (Country.DoesNotExist, State.DoesNotExist):

                self.stdout.write(
                    self.style.ERROR(
                        f"Cannot import {city_name}"
                    )
                )

        self.stdout.write(
            self.style.SUCCESS(
                "All Locations Imported Successfully."
            )
        )