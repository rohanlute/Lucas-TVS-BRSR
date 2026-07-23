"""
excel_template_service.py

Two jobs:
  1. build_scope_template_workbook()  -> in-memory .xlsx the user downloads,
     fills in offline, and re-uploads.
  2. parse_scope_template_workbook()  -> reads that filled-in file back into
     plain Python dicts your existing /api/save-transactions/ logic can save.
"""

import io
import re
from decimal import Decimal
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

# Import models from your app
from apps.companies.models import Company
from apps.organizations.models import (
    FinancialYear,
    FinancialMonth,
    Plant,
)
from apps.calculator.models import Unit

# Import the emission models
from .models import (
    EmissionScope,
    EmissionCategory,
    EmissionActivity,
    EmissionSource,
    EmissionFactor,
    EmissionAssignment,
    EmissionTransaction,
)

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
EDIT_FILL = PatternFill("solid", fgColor="FFF9DB")   # yellow = "fill this in"
LOCK_FILL = PatternFill("solid", fgColor="F1F5F9")   # grey = locked/computed
THIN = Side(style="thin", color="D0D7E2")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
SHEET_NAME_BAD_CHARS = re.compile(r"[:\\/?*\[\]]")


def _safe_sheet_name(name, used):
    """Create a safe Excel sheet name that doesn't exceed 31 chars."""
    name = SHEET_NAME_BAD_CHARS.sub("-", name)[:31] or "Category"
    base, i = name, 2
    while name in used:
        suffix = f" ({i})"
        name = base[: 31 - len(suffix)] + suffix
        i += 1
    used.add(name)
    return name


def _lock_sheet(ws, editable_cells):
    """Lock every cell except the ones explicitly passed in."""
    editable = set(editable_cells)
    for row in ws.iter_rows():
        for cell in row:
            cell.protection = cell.protection.copy(
                locked=(cell.coordinate not in editable)
            )
    ws.protection.sheet = True
    # Set empty password instead of None to avoid openpyxl error
    ws.protection.password = ""


def build_scope_template_workbook(scope, categories, company, plants,
                                   financial_years, financial_months,
                                   assignment=None):
    """
    Build an Excel template for scope data entry.
    
    Args:
        scope: EmissionScope object
        categories: Iterable of EmissionCategory objects with activities_list
        company: Company object (auto-filled, LOCKED)
        plants: QuerySet/list of Plant objects (for dropdown - EDITABLE)
        financial_years: QuerySet/list of FinancialYear objects (for dropdown - EDITABLE)
        financial_months: QuerySet/list of FinancialMonth objects (for dropdown - EDITABLE)
        assignment: Optional EmissionAssignment to preselect values
    
    Returns:
        tuple: (BytesIO, filename)
    """
    wb = Workbook()

    # ---------------- hidden "Lists" sheet (data validation source) --------
    lists_ws = wb.active
    lists_ws.title = "Lists"
    lists_ws.sheet_state = "hidden"

    plant_names = [p.name for p in plants]
    fy_names = [str(fy.financial_year) for fy in financial_years]
    month_names = [m.month_name for m in financial_months]

    for col, values in zip("ABC", (plant_names, fy_names, month_names)):
        for i, v in enumerate(values, start=1):
            lists_ws[f"{col}{i}"] = v

    def _list_range(col, count):
        return f"Lists!${col}$1:${col}${max(count, 1)}"

    # ---------------- Context Info sheet ------------------------------------
    ctx = wb.create_sheet("Context Info", 0)
    ctx.sheet_view.showGridLines = False
    ctx.column_dimensions["A"].width = 26
    ctx.column_dimensions["B"].width = 40

    ctx["A1"] = "SCOPE DATA ENTRY — CONTEXT INFORMATION"
    ctx["A1"].font = Font(bold=True, size=13, color="1F2937")
    ctx.merge_cells("A1:B1")

    ctx["A2"] = f"Scope: {scope.name}"
    ctx["A2"].font = Font(italic=True, color="6B7280")
    ctx.merge_cells("A2:B2")

    # Preselect values from assignment if provided
    default_plant = assignment.plant.name if assignment and assignment.plant else (
        plants[0].name if plants else ""
    )
    default_fy = str(assignment.financial_year.financial_year) if assignment and assignment.financial_year else (
        str(financial_years[0].financial_year) if financial_years else ""
    )
    default_month = assignment.financial_month.month_name if assignment and assignment.financial_month else (
        financial_months[0].month_name if financial_months else ""
    )

    # ============================================================
    # KEY CHANGE: Company is LOCKED (editable=False), 
    # Plant, Financial Year, Financial Month are EDITABLE (editable=True)
    # ============================================================
    rows = [
        ("Company", company.company_name, False, None),  # LOCKED - auto-filled
        ("Plant", default_plant, True, _list_range("A", len(plant_names))),  # EDITABLE - dropdown
        ("Financial Year", default_fy, True, _list_range("B", len(fy_names))),  # EDITABLE - dropdown
        ("Financial Month", default_month, True, _list_range("C", len(month_names))),  # EDITABLE - dropdown
    ]

    start_row = 4
    editable_cells = []
    for i, (label, value, editable, dv_range) in enumerate(rows):
        r = start_row + i
        ctx[f"A{r}"] = label
        ctx[f"A{r}"].font = Font(bold=True, color="374151")
        cell = ctx[f"B{r}"]
        cell.value = value
        cell.fill = EDIT_FILL if editable else LOCK_FILL  # Yellow for editable, Grey for locked
        cell.border = BORDER
        if editable:
            editable_cells.append(cell.coordinate)
            dv = DataValidation(type="list", formula1=f"={dv_range}",
                                 allow_blank=False, showDropDown=False)
            ctx.add_data_validation(dv)
            dv.add(cell)
        else:
            # Company is locked - make it read-only
            cell.protection = cell.protection.copy(locked=True)

    # Add a helpful note
    legend_row = start_row + len(rows) + 2
    ctx[f"A{legend_row}"] = "Yellow cells = choose from dropdown. Grey cells are locked/auto-filled."
    ctx[f"A{legend_row}"].font = Font(italic=True, size=10, color="9CA3AF")
    ctx.merge_cells(f"A{legend_row}:B{legend_row}")

    # Add instruction for Company
    instruction_row = start_row + len(rows) + 3
    ctx[f"A{instruction_row}"] = "Note: Company is auto-filled based on your account. You can only edit Plant, Financial Year, and Financial Month."
    ctx[f"A{instruction_row}"].font = Font(italic=True, size=9, color="6B7280")
    ctx.merge_cells(f"A{instruction_row}:B{instruction_row}")

    _lock_sheet(ctx, editable_cells)

    # ---------------- one sheet per category --------------------------------
    used_names = {"Context Info", "Lists"}
    headers = ["ACTIVITY", "SOURCE", "QUANTITY", "UNIT",
               "EMISSION FACTOR", "TOTAL EMISSION (tCO2e)",
               "activity_id", "source_id", "unit_id", "category_id"]

    for category in categories:
        sheet_name = _safe_sheet_name(category.name, used_names)
        ws = wb.create_sheet(sheet_name)
        ws.sheet_view.showGridLines = False

        ws.merge_cells("A1:F1")
        ws["A1"] = category.name
        ws["A1"].font = Font(bold=True, size=13, color="1F2937")
        
        # Add category description if available
        if category.description:
            ws["A2"] = category.description
            ws["A2"].font = Font(italic=True, size=10, color="6B7280")
            ws.merge_cells("A2:F2")

        header_row = 3
        for c, h in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=c, value=h)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
            cell.border = BORDER

        editable_cells = []
        r = header_row + 1
        
        # Use activities_list instead of activities (custom attribute)
        activities = getattr(category, 'activities_list', [])
        
        for activity in activities:
            # Use sources_list instead of sources (custom attribute)
            sources = getattr(activity, 'sources_list', [])
            if not sources:
                # If no sources, create a row with empty source
                sources = [None]
                
            for source in sources:
                # Get emission factor for this activity/source
                factor = Decimal('0')
                if hasattr(activity, 'factor_for') and callable(activity.factor_for):
                    factor = activity.factor_for(source)
                elif source and activity.requires_emission_factor:
                    # Fallback: try to get factor from EmissionFactor
                    factor_obj = EmissionFactor.objects.filter(
                        activity=activity,
                        is_active=True,
                        effective_from__lte=timezone.now().date()
                    ).order_by("-effective_from").first()
                    if factor_obj:
                        factor = factor_obj.emission_factor

                # Activity name (locked)
                ws.cell(row=r, column=1, value=activity.name).border = BORDER
                
                # Source name (locked)
                source_name = source.source_name if source else "No Source"
                ws.cell(row=r, column=2, value=source_name).border = BORDER

                # Quantity cell (EDITABLE - yellow)
                qty_cell = ws.cell(row=r, column=3, value=None)
                qty_cell.fill = EDIT_FILL
                qty_cell.border = BORDER
                editable_cells.append(qty_cell.coordinate)

                # Unit name (locked)
                unit_name = activity.base_unit.symbol if activity.base_unit else "Unit"
                ws.cell(row=r, column=4, value=unit_name).border = BORDER

                # Emission factor (locked - grey)
                factor_cell = ws.cell(row=r, column=5, value=float(factor))
                factor_cell.fill = LOCK_FILL
                factor_cell.border = BORDER

                # Total emission formula (locked - grey)
                col_letter_qty = get_column_letter(3)
                col_letter_factor = get_column_letter(5)
                total_cell = ws.cell(
                    row=r, column=6,
                    value=f"=IF({col_letter_qty}{r}=\"\",0,"
                          f"{col_letter_qty}{r}*{col_letter_factor}{r})"
                )
                total_cell.fill = LOCK_FILL
                total_cell.border = BORDER

                # Hidden mapping columns, used on re-import
                ws.cell(row=r, column=7, value=activity.id)
                ws.cell(row=r, column=8, value=(source.id if source else ""))
                ws.cell(row=r, column=9, value=(activity.base_unit.id if activity.base_unit else ""))
                ws.cell(row=r, column=10, value=category.id)

                r += 1

        # Hide the mapping columns
        for col in "GHIJ":
            ws.column_dimensions[col].hidden = True
        for col, width in zip("ABCDEF", (30, 24, 14, 12, 16, 20)):
            ws.column_dimensions[col].width = width
        ws.freeze_panes = "A4"

        _lock_sheet(ws, editable_cells)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"scope_{scope.name.replace(' ', '_').lower()}_template.xlsx"
    return buf, filename


def parse_scope_template_workbook(file_obj):
    """
    Parse an uploaded filled-in template back into plain data.
    
    Args:
        file_obj: Uploaded file object
    
    Returns:
        dict: {
            "plant_name": str,
            "financial_year": str,
            "financial_month": str,
            "rows": [
                {
                    "activity_id": int,
                    "source_id": int|None,
                    "unit_id": int,
                    "category_id": int,
                    "quantity": float,
                    "factor": float,
                    "total": float
                },
                ...
            ],
            "errors": [str, ...]
        }
    
    Raises:
        ValueError: On structurally invalid files
    """
    wb = load_workbook(file_obj, data_only=True)

    if "Context Info" not in wb.sheetnames:
        raise ValueError("This doesn't look like a scope template — "
                         "'Context Info' sheet is missing.")

    ctx = wb["Context Info"]
    context = {}
    
    # Parse context info (rows 4-7)
    # Row 4: Company (locked, we ignore it)
    # Row 5: Plant (editable)
    # Row 6: Financial Year (editable)
    # Row 7: Financial Month (editable)
    for row in ctx.iter_rows(min_row=4, max_row=7, max_col=2):
        label_cell, value_cell = row[0], row[1]
        if not label_cell.value:
            continue
        label = str(label_cell.value).strip().lower()
        if label == "plant":
            context["plant_name"] = value_cell.value
        elif label == "financial year":
            context["financial_year"] = value_cell.value
        elif label == "financial month":
            context["financial_month"] = value_cell.value

    missing = [k for k in ("plant_name", "financial_year", "financial_month")
               if not context.get(k)]
    if missing:
        raise ValueError(f"Context Info sheet is missing: {', '.join(missing)}")

    rows_out = []
    errors = []
    reserved = {"Context Info", "Lists"}
    
    for sheet_name in wb.sheetnames:
        if sheet_name in reserved:
            continue
        ws = wb[sheet_name]
        header_row = 3
        
        for r in range(header_row + 1, ws.max_row + 1):
            activity_id = ws.cell(row=r, column=7).value
            if not activity_id:
                continue
                
            qty = ws.cell(row=r, column=3).value
            try:
                qty = float(qty) if qty not in (None, "") else 0.0
            except (TypeError, ValueError):
                errors.append(f"{sheet_name} row {r}: quantity '{qty}' is not "
                              f"a number, row skipped.")
                continue
                
            if qty <= 0:
                continue

            factor = ws.cell(row=r, column=5).value or 0
            total = ws.cell(row=r, column=6).value
            if total in (None, ""):
                total = qty * float(factor)

            source_id = ws.cell(row=r, column=8).value
            unit_id = ws.cell(row=r, column=9).value
            
            if not unit_id:
                errors.append(f"{sheet_name} row {r}: missing unit ID, row skipped.")
                continue

            rows_out.append({
                "activity_id": int(activity_id),
                "source_id": int(source_id) if source_id not in (None, "") else None,
                "unit_id": int(unit_id),
                "category_id": int(ws.cell(row=r, column=10).value),
                "quantity": qty,
                "factor": float(factor),
                "total": float(total),
            })

    context["rows"] = rows_out
    context["errors"] = errors
    return context